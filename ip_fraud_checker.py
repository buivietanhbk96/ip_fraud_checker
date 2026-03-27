"""
IP Fraud Score Checker Tool
===========================
Kiểm tra fraud score của các IP trong một dải subnet cấu hình được từ ip2location.com

Hỗ trợ 2 phương pháp:
    1. API ip2location.io (nhanh, ổn định - cần free API key)
    2. Web scraping trang demo (không cần API key, chậm hơn)

Đăng ký free API key tại: https://www.ip2location.io/sign-up
    -> Free plan: 200 lookups/ngày (đủ cho nhu cầu 10-20 IP)

Cách dùng:
    python ip_fraud_checker.py                          # Web scraping, 10-20 random IPs
    python ip_fraud_checker.py --api-key YOUR_KEY      # Dùng API
    python ip_fraud_checker.py --count 15              # Chỉ định số IP cụ thể
    python ip_fraud_checker.py --all                   # Check tất cả IP host (chỉ nên dùng với API)
    python ip_fraud_checker.py --subnet 10.10.10.0/24  # Dải IP tùy chỉnh
    python ip_fraud_checker.py --schedule 60           # Tự động chạy mỗi 60 phút
"""

import argparse
import ipaddress
import json
import logging
import os
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# Configuration
# ============================================================================

DEFAULT_SUBNET = "103.94.16.0/24"
DEFAULT_MIN_IPS = 10
DEFAULT_MAX_IPS = 20
CONFIG_FILE = Path(__file__).parent / "config.json"

# API endpoint
API_URL = "https://api.ip2location.io"

# Web scraping endpoint
DEMO_URL = "https://www.ip2location.com/demo/{ip}"

# Request settings
REQUEST_TIMEOUT = 30
SCRAPE_DELAY_MIN = 2  # Delay giữa các request khi scraping (giây)
SCRAPE_DELAY_MAX = 5
API_DELAY = 0.5  # Delay giữa các API call (giây)

# User-Agent để tránh bị block
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/131.0.0.0 Safari/537.36"
)

# Output directory
OUTPUT_DIR = Path(__file__).parent / "output"

# Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def load_config(config_path: Path = CONFIG_FILE) -> dict:
    """Đọc cấu hình từ file JSON nếu có."""
    if not config_path.exists():
        logger.info(
            f"Không tìm thấy file cấu hình {config_path.name}, dùng cấu hình mặc định trong code"
        )
        return {}

    try:
        with config_path.open("r", encoding="utf-8") as file:
            data = json.load(file)

        if not isinstance(data, dict):
            logger.warning(
                f"File {config_path.name} không đúng định dạng object JSON, bỏ qua cấu hình"
            )
            return {}

        logger.info(f"Đã nạp cấu hình từ {config_path.name}")
        return data
    except json.JSONDecodeError as error:
        logger.warning(
            f"Không đọc được {config_path.name} do JSON không hợp lệ: {error}. Dùng mặc định trong code"
        )
        return {}
    except OSError as error:
        logger.warning(
            f"Không đọc được {config_path.name}: {error}. Dùng mặc định trong code"
        )
        return {}


def resolve_subnet(cli_subnet: str | None, config: dict | None = None) -> str:
    """Xác định subnet theo thứ tự ưu tiên: CLI > config.json > code."""
    if cli_subnet:
        return cli_subnet

    if config:
        configured_subnet = config.get("default_subnet")
        if isinstance(configured_subnet, str) and configured_subnet.strip():
            return configured_subnet.strip()

    return DEFAULT_SUBNET


def validate_subnet(subnet: str) -> str:
    """Validate subnet và trả về chuỗi chuẩn hóa."""
    return str(ipaddress.ip_network(subnet, strict=False))


# ============================================================================
# IP Generation
# ============================================================================


def get_all_ips(subnet: str = DEFAULT_SUBNET) -> list[str]:
    """Lấy tất cả IP host trong dải subnet."""
    network = ipaddress.ip_network(subnet, strict=False)
    return [str(ip) for ip in network.hosts()]


def get_random_ips(
    count: int | None = None, subnet: str = DEFAULT_SUBNET
) -> list[str]:
    """Random một số lượng IP từ dải subnet."""
    all_ips = get_all_ips(subnet)

    if count is None:
        count = random.randint(DEFAULT_MIN_IPS, DEFAULT_MAX_IPS)

    count = min(count, len(all_ips))
    selected = random.sample(all_ips, count)
    selected.sort(key=lambda ip: ipaddress.ip_address(ip))

    logger.info(f"Đã chọn ngẫu nhiên {len(selected)} IP từ dải {subnet}")
    return selected


# ============================================================================
# Method 1: API ip2location.io (Recommended)
# ============================================================================


def check_ip_via_api(ip: str, api_key: str) -> dict | None:
    """Kiểm tra fraud score của 1 IP qua API ip2location.io."""
    try:
        params = {
            "key": api_key,
            "ip": ip,
            "format": "json",
        }
        response = requests.get(API_URL, params=params, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()

        data = response.json()

        if "error" in data:
            logger.error(f"API error cho {ip}: {data['error']}")
            return None

        result = {
            "ip": ip,
            "fraud_score": data.get("fraud_score", "N/A"),
            "is_proxy": data.get("is_proxy", "N/A"),
            "country": data.get("country_name", "N/A"),
            "region": data.get("region_name", "N/A"),
            "city": data.get("city_name", "N/A"),
            "isp": data.get("isp", "N/A"),
            "usage_type": data.get("usage_type", "N/A"),
            "proxy_type": data.get("proxy", {}).get("proxy_type", "N/A"),
            "threat": data.get("proxy", {}).get("threat", "N/A"),
            "is_vpn": data.get("proxy", {}).get("is_vpn", "N/A"),
            "is_tor": data.get("proxy", {}).get("is_tor", "N/A"),
            "is_data_center": data.get("proxy", {}).get("is_data_center", "N/A"),
            "is_public_proxy": data.get("proxy", {}).get("is_public_proxy", "N/A"),
            "method": "API",
        }

        logger.info(
            f"[API] {ip} -> Fraud Score: {result['fraud_score']}, "
            f"Proxy: {result['is_proxy']}"
        )
        return result

    except requests.exceptions.RequestException as e:
        logger.error(f"Request error cho {ip}: {e}")
        return None
    except (json.JSONDecodeError, KeyError) as e:
        logger.error(f"Parse error cho {ip}: {e}")
        return None


# ============================================================================
# Method 2: Web Scraping (Fallback - No API key needed)
# ============================================================================


def check_ip_via_scraping(ip: str, session: requests.Session) -> dict | None:
    """Kiểm tra fraud score của 1 IP bằng web scraping trang demo.

    Cấu trúc HTML trang demo ip2location:
    - Table 0 = Geolocation Data (mỗi row: <th>label</th><td>value</td>)
    - Table 1 = Proxy Data (mỗi row: <th>label</th><td>value</td>)
    - Fraud Score nằm ở row cuối Table 1
    - Trang cũng embed JSON ví dụ API có chứa "fraud_score": N
    """
    url = DEMO_URL.format(ip=ip)

    try:
        response = session.get(url, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()

        html = response.text
        soup = BeautifulSoup(html, "lxml")

        result = {
            "ip": ip,
            "fraud_score": "N/A",
            "is_proxy": "N/A",
            "country": "N/A",
            "region": "N/A",
            "city": "N/A",
            "isp": "N/A",
            "usage_type": "N/A",
            "proxy_type": "N/A",
            "threat": "N/A",
            "is_vpn": "N/A",
            "is_tor": "N/A",
            "is_data_center": "N/A",
            "is_public_proxy": "N/A",
            "method": "Scraping",
        }

        tables = soup.find_all("table")

        # --- Parse Table 0: Geolocation Data ---
        # Mỗi row có <th>Label</th> <td>Value</td> hoặc chỉ <td>Value</td>
        geo_field_map = {
            "country": "country",
            "region": "region",
            "city": "city",
            "isp": "isp",
            "usage type": "usage_type",
        }

        if len(tables) >= 1:
            for row in tables[0].find_all("tr"):
                th = row.find("th")
                td = row.find("td")
                if th and td:
                    label = th.get_text(strip=True).lower()
                    value = td.get_text(strip=True)
                    for pattern, key in geo_field_map.items():
                        if pattern in label:
                            result[key] = value
                            break

        # --- Parse Table 1: Proxy Data ---
        # Row order: IP, Anonymous Proxy, Proxy Country, ..., Fraud Score (last)
        proxy_field_map = {
            "anonymous proxy": "is_proxy",
            "proxy type": "proxy_type",
            "threat": "threat",
            "fraud score": "fraud_score",
        }

        if len(tables) >= 2:
            for row in tables[1].find_all("tr"):
                th = row.find("th")
                td = row.find("td")
                if th and td:
                    label = th.get_text(strip=True).lower()
                    value = td.get_text(strip=True)
                    for pattern, key in proxy_field_map.items():
                        if pattern in label:
                            result[key] = value
                            break

            # Fallback: nếu không tìm thấy qua th/td, lấy row cuối của Table 1
            if result["fraud_score"] == "N/A":
                proxy_rows = tables[1].find_all("tr")
                if proxy_rows:
                    last_td = proxy_rows[-1].find("td")
                    if last_td:
                        val = last_td.get_text(strip=True)
                        if val.isdigit():
                            result["fraud_score"] = val

        # Fallback 2: tìm "fraud_score": N trong JSON embedded trên trang
        if result["fraud_score"] == "N/A":
            match = re.search(r'"fraud_score"\s*:\s*(\d+)', html)
            if match:
                result["fraud_score"] = match.group(1)

        # Fallback 3: tìm text "Fraud Score" + số
        if result["fraud_score"] == "N/A":
            text = soup.get_text()
            match = re.search(
                r"Fraud\s*Score\s*[:\|]?\s*(\d+)", text, re.IGNORECASE
            )
            if match:
                result["fraud_score"] = match.group(1)

        logger.info(
            f"[Scraping] {ip} -> Fraud Score: {result['fraud_score']}, "
            f"Proxy: {result['is_proxy']}"
        )
        return result

    except requests.exceptions.RequestException as e:
        logger.error(f"Request error cho {ip}: {e}")
        return None
    except Exception as e:
        logger.error(f"Parse error cho {ip}: {e}")
        return None


# ============================================================================
# Main Check Logic
# ============================================================================


def check_ips(ip_list: list[str], api_key: str | None = None) -> list[dict]:
    """Kiểm tra fraud score cho danh sách IP."""
    results = []
    total = len(ip_list)

    if api_key:
        logger.info(f"Sử dụng API mode với {total} IPs")
        for i, ip in enumerate(ip_list, 1):
            logger.info(f"[{i}/{total}] Đang kiểm tra {ip}...")
            result = check_ip_via_api(ip, api_key)
            if result:
                results.append(result)
            if i < total:
                time.sleep(API_DELAY)
    else:
        logger.info(f"Sử dụng Web Scraping mode với {total} IPs")
        logger.warning(
            "⚠ Web scraping có thể chậm hơn và bị rate-limit. "
            "Nên dùng API key để ổn định hơn."
        )
        session = requests.Session()
        session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
                "Connection": "keep-alive",
            }
        )

        for i, ip in enumerate(ip_list, 1):
            logger.info(f"[{i}/{total}] Đang kiểm tra {ip}...")
            result = check_ip_via_scraping(ip, session)
            if result:
                results.append(result)

            if i < total:
                delay = random.uniform(SCRAPE_DELAY_MIN, SCRAPE_DELAY_MAX)
                logger.debug(f"  Đợi {delay:.1f}s trước request tiếp...")
                time.sleep(delay)

    logger.info(f"Hoàn thành: {len(results)}/{total} IP đã kiểm tra thành công")
    return results


# ============================================================================
# Excel Export
# ============================================================================


def export_to_excel(
    results: list[dict], subnet: str, output_path: Path | None = None
) -> Path:
    """Xuất kết quả ra file Excel với formatting đẹp."""
    if not results:
        logger.warning("Không có kết quả để xuất!")
        return None

    # Tạo output directory nếu chưa có
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_DIR / f"ip_fraud_scores_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "IP Fraud Scores"

    # ---- Styles ----
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(horizontal="center", vertical="center")
    data_alignment_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Fraud score color fills
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
    red_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")

    # ---- Title row ----
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"IP Fraud Score Report - Dải {subnet} - "
        f"Ngày: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    )
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ---- Headers ----
    headers = [
        ("STT", 6),
        ("IP Address", 18),
        ("Fraud Score", 14),
        ("Is Proxy", 11),
        ("Country", 22),
        ("Region", 18),
        ("City", 16),
        ("ISP", 25),
        ("Usage Type", 14),
        ("Proxy Type", 14),
        ("Threat", 14),
        ("Is VPN", 10),
        ("Is TOR", 10),
        ("Is Data Center", 15),
        ("Method", 10),
    ]

    for col_idx, (header_name, width) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 25

    # ---- Data rows ----
    for row_idx, result in enumerate(results, 3):
        data_row = [
            row_idx - 2,  # STT
            result["ip"],
            result["fraud_score"],
            result["is_proxy"],
            result["country"],
            result["region"],
            result["city"],
            result["isp"],
            result["usage_type"],
            result["proxy_type"],
            result["threat"],
            result.get("is_vpn", "N/A"),
            result.get("is_tor", "N/A"),
            result.get("is_data_center", "N/A"),
            result["method"],
        ]

        for col_idx, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            if col_idx in (1, 3, 4, 9, 10, 11, 12, 13, 14, 15):
                cell.alignment = data_alignment
            else:
                cell.alignment = data_alignment_left

        # Color-code fraud score
        fraud_cell = ws.cell(row=row_idx, column=3)
        try:
            score = int(str(result["fraud_score"]).strip())
            if score <= 20:
                fraud_cell.fill = green_fill
            elif score <= 50:
                fraud_cell.fill = yellow_fill
            elif score <= 80:
                fraud_cell.fill = orange_fill
            else:
                fraud_cell.fill = red_fill
                fraud_cell.font = red_font
        except (ValueError, TypeError):
            pass

    # ---- Summary sheet ----
    ws_summary = wb.create_sheet("Summary")
    ws_summary.merge_cells("A1:D1")
    ws_summary["A1"].value = "Tổng hợp Fraud Score"
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    ws_summary["A1"].alignment = Alignment(horizontal="center")

    # Summary headers
    summary_headers = ["Metric", "Value", "Details", ""]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Calculate stats
    scores = []
    for r in results:
        try:
            scores.append(int(str(r["fraud_score"]).strip()))
        except (ValueError, TypeError):
            pass

    summary_data = [
        ("Tổng số IP kiểm tra", len(results), ""),
        ("IP lấy thành công fraud score", len(scores), ""),
        ("Dải IP", subnet, ""),
        ("Thời gian kiểm tra", datetime.now().strftime("%d/%m/%Y %H:%M:%S"), ""),
        ("", "", ""),
        ("--- Fraud Score Statistics ---", "", ""),
        ("Fraud Score thấp nhất", min(scores) if scores else "N/A", ""),
        ("Fraud Score cao nhất", max(scores) if scores else "N/A", ""),
        (
            "Fraud Score trung bình",
            f"{sum(scores) / len(scores):.1f}" if scores else "N/A",
            "",
        ),
        ("", "", ""),
        ("--- Phân loại ---", "", ""),
        (
            "Score 0-20 (Low Risk)",
            sum(1 for s in scores if s <= 20),
            "🟢 An toàn",
        ),
        (
            "Score 21-50 (Medium Risk)",
            sum(1 for s in scores if 21 <= s <= 50),
            "🟡 Cần chú ý",
        ),
        (
            "Score 51-80 (High Risk)",
            sum(1 for s in scores if 51 <= s <= 80),
            "🟠 Nguy hiểm",
        ),
        (
            "Score 81-100 (Critical)",
            sum(1 for s in scores if s > 80),
            "🔴 Rất nguy hiểm",
        ),
    ]

    for row_idx, (metric, value, detail) in enumerate(summary_data, 3):
        ws_summary.cell(row=row_idx, column=1, value=metric).font = data_font
        ws_summary.cell(row=row_idx, column=2, value=value).font = Font(
            name="Calibri", bold=True, size=11
        )
        ws_summary.cell(row=row_idx, column=3, value=detail).font = data_font

        for col_idx in range(1, 4):
            ws_summary.cell(row=row_idx, column=col_idx).border = thin_border

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 25
    ws_summary.column_dimensions["C"].width = 20

    # ---- Auto-filter ----
    ws.auto_filter.ref = f"A2:O{len(results) + 2}"

    # ---- Freeze panes ----
    ws.freeze_panes = "A3"

    # ---- Save ----
    wb.save(output_path)
    logger.info(f"✅ Đã xuất kết quả ra: {output_path}")
    logger.info(f"   Tổng số IP: {len(results)}")
    if scores:
        logger.info(f"   Fraud Score range: {min(scores)} - {max(scores)}")
        logger.info(f"   Fraud Score trung bình: {sum(scores) / len(scores):.1f}")

    return output_path


# ============================================================================
# Scheduler
# ============================================================================


def run_scheduled(interval_minutes: int, **kwargs):
    """Chạy kiểm tra định kỳ."""
    logger.info(f"🕐 Bắt đầu chế độ scheduled - chạy mỗi {interval_minutes} phút")
    logger.info("   Nhấn Ctrl+C để dừng")

    run_count = 0
    while True:
        run_count += 1
        logger.info(f"\n{'='*60}")
        logger.info(f"Lần chạy #{run_count} - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        logger.info(f"{'='*60}")

        try:
            run_once(**kwargs)
        except Exception as e:
            logger.error(f"Lỗi trong lần chạy #{run_count}: {e}")

        logger.info(
            f"Đợi {interval_minutes} phút đến lần chạy tiếp theo... "
            f"(Ctrl+C để dừng)"
        )
        time.sleep(interval_minutes * 60)


def run_once(
    subnet: str = DEFAULT_SUBNET,
    api_key: str | None = None,
    count: int | None = None,
    check_all: bool = False,
    output_path: Path | None = None,
) -> Path | None:
    """Chạy kiểm tra 1 lần."""
    logger.info(f"🚀 Bắt đầu kiểm tra fraud score cho dải {subnet}")

    # Chọn IP
    if check_all:
        ip_list = get_all_ips(subnet)
        logger.info(f"Chế độ: Kiểm tra TẤT CẢ {len(ip_list)} IP")
    else:
        ip_list = get_random_ips(count, subnet)
        logger.info(
            f"Chế độ: Random {len(ip_list)} IP "
            f"(từ {DEFAULT_MIN_IPS}-{DEFAULT_MAX_IPS})"
        )

    # Kiểm tra
    results = check_ips(ip_list, api_key)

    # Xuất Excel
    if results:
        return export_to_excel(results, subnet, output_path)
    else:
        logger.error("Không có kết quả nào! Kiểm tra lại kết nối mạng hoặc API key.")
        return None


# ============================================================================
# CLI
# ============================================================================


def parse_args():
    parser = argparse.ArgumentParser(
        description="IP Fraud Score Checker - Kiểm tra fraud score từ ip2location.com",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ví dụ:
  python ip_fraud_checker.py                          # Scraping, 10-20 random IPs
  python ip_fraud_checker.py --api-key YOUR_KEY       # Dùng API (khuyến nghị)
  python ip_fraud_checker.py --count 15               # Chỉ định số lượng IP
    python ip_fraud_checker.py --all --api-key KEY      # Check tất cả IP host
  python ip_fraud_checker.py --schedule 60            # Chạy mỗi 60 phút
  python ip_fraud_checker.py --subnet 103.94.16.0/24  # Dải IP tùy chỉnh

Đăng ký free API key: https://www.ip2location.io/sign-up
  -> Free plan cho 200 lookups/ngày (đủ cho nhu cầu)
        """,
    )

    parser.add_argument(
        "--api-key",
        type=str,
        default=os.environ.get("IP2LOCATION_API_KEY"),
        help="API key của ip2location.io (hoặc set env IP2LOCATION_API_KEY)",
    )
    parser.add_argument(
        "--count",
        type=int,
        default=None,
        help=f"Số lượng IP random (mặc định: {DEFAULT_MIN_IPS}-{DEFAULT_MAX_IPS})",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Kiểm tra tất cả IP trong dải (cần API key)",
    )
    parser.add_argument(
        "--subnet",
        type=str,
        default=None,
        help=(
            "Dải IP cần kiểm tra. Độ ưu tiên: --subnet > config.json > "
            f"DEFAULT_SUBNET trong code ({DEFAULT_SUBNET})"
        ),
    )
    parser.add_argument(
        "--schedule",
        type=int,
        default=None,
        help="Chạy định kỳ mỗi N phút",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Đường dẫn file Excel output",
    )

    return parser.parse_args()


def main():
    args = parse_args()

    config = load_config()
    try:
        subnet = validate_subnet(resolve_subnet(args.subnet, config))
    except ValueError as error:
        logger.error(f"Dải IP không hợp lệ: {error}")
        sys.exit(1)

    logger.info(f"Dải IP đang dùng: {subnet}")

    # Validate
    if args.all and not args.api_key:
        logger.warning(
            "⚠ Check tất cả IP bằng web scraping sẽ rất chậm và có thể bị block. "
            "Khuyến nghị dùng --api-key"
        )

    output_path = Path(args.output) if args.output else None

    kwargs = {
        "subnet": subnet,
        "api_key": args.api_key,
        "count": args.count,
        "check_all": args.all,
        "output_path": output_path,
    }

    if args.schedule:
        run_scheduled(args.schedule, **kwargs)
    else:
        result_path = run_once(**kwargs)
        if result_path:
            print(f"\n{'='*60}")
            print(f"✅ File kết quả: {result_path}")
            print(f"{'='*60}")


if __name__ == "__main__":
    main()
